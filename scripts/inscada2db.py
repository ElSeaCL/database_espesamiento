import numpy as np
from openpyxl import load_workbook
from pathlib import Path, PureWindowsPath

# archivo donde se guardará
excel_db_file = PureWindowsPath("T:/PROCESOS/18. Seguimientos/Espesamiento/db/excel/db_espesamiento_prueba.xlsx")
inscada_file = PureWindowsPath("T:/BALANCE/IN SCADA/Inscada 2020.xlsm")

inscada = load_workbook(filename = inscada_file, data_only=True) 
excel_db = load_workbook(filename = excel_db_file, data_only=True)

def array2list(array):
    '''
        Devuelve una lista a partir de un array de openpyxl 
    '''

    return [x[0].value for x in array]

def xl2npMatrix(hoja, minrow, maxrow, mincol, maxcol):
    '''
        Toma una hoja de excel y tras definir los limites de extracción de datos,
        (fila minima, fila máxima, columna minima, columna máxima) devuelve una
        matriz de numpy.

        :param hoja: Nombre de la hoja de excel de la que extraen los datos
        :param minrow: Numero de la fila donde se empieza a rescatar los valores.
        :param maxrow: Numero de la fila donde termina de rescatar los valores.
        :param mincol: Numero de la columna donde se empieza a rescatar los valores.
        :param maxcol: Numero de la columna donde termina de rescatar los valores.
        :returns: Matriz de Numpy con los valores.

        >>> xl2npMatrix('hoja1', 1, 3, 2, 4)
        array([[1, 0, 1],
               [1, 0, 1]])
    '''

    listo = list()

    for row in hoja.iter_rows(min_row = minrow, max_row = maxrow, 
    min_col = mincol, max_col = maxcol):
        for cell in row:
            listo.append(cell.value)
    
    listo = np.array(listo)
    listo = listo.reshape(maxrow - minrow + 1, maxcol - mincol + 1)

    return listo

def splitnonalpha(s):
   '''
        Toma como argumento una cadena y la separa en sus secciones alfabeticas.
   ''' 

   pos = 1
   while pos < len(s) and s[pos].isalpha():
      pos+=1
   return (s[:pos], s[pos:])

def matrix2excel(hoja, celda, matriz):
    '''
        Toma una matriz (numpy.matrix) y la inserta en la hoja seleccionada a
        partir de la celda elegida.

        :param hoja: Hoja donde se almacenaran los datos. Debe ser un Worksheet de openpyxl
        :param celda: Celda donde se comenzará a almacenar lso datos. Esquina superior isquierda de la matriz
        :param matriz: Matriz a grabar
    '''

    # Arreglo de la matríz
    fila, col = matriz.shape

    # Transformación de la celda, de formato escrito a numerico
    celda_col, celda_fila = splitnonalpha(celda)
    num_fila = int(celda_fila) 

    # Se transforma la cadena alfabetica en un indicador numerico de la columna 
    pos = 1
    num_col = 0
    while pos <= len(celda_col):
        num_col += (ord(celda_col[-pos])-64)*pow(26,(pos-1))
        pos += 1

    # Alamacenamiento de los valores segun la hoja seleccionada y la celda de referencia
    cont_col = 0
    cont_fila = 0
 
    for i in range(fila*col):
        hoja.cell(num_fila + cont_fila, num_col + cont_col).value = matriz.item(i)
        cont_col += 1

        if cont_col % col == 0:
            cont_col = 0
            cont_fila += 1

####################################################################################################################
# MAIN
####################################################################################################################

# Carga la hoja a utilizar
espesamiento = inscada['Espesamiento 2°']

# Rescate de valores

# array de zeros
zeros = np.zeros((31,2), dtype='float')

## Matriz con los valores de lodo
lodos = xl2npMatrix(espesamiento, 101, 131, 9, 16)

## MAtriz con los valores de polimero
polimeros = xl2npMatrix(espesamiento, 101, 131, 23, 30)

## Matriz con los valores de torque
torque = xl2npMatrix(espesamiento, 101, 131, 31, 36)
for i in range(31):
    for j in range(6):
        if torque[i,j][-1].isalpha() == True:
            torque[i,j] = '0'
torque = torque.astype('float')
torque = np.concatenate((torque,zeros), axis=1)

## Matriz con los valores de vr
vr = xl2npMatrix(espesamiento, 101, 131, 37, 42)
for i in range(31):
    for j in range(6):
        if vr[i,j][-1].isalpha() == True:
            vr[i,j] = '0'
vr = vr.astype('float')
vr = np.concatenate((vr, zeros), axis=1)

# Reordenamos las matrices
lodos = lodos.reshape((31*8,1))
polimeros = polimeros.reshape((31*8,1))
torque = torque.reshape((31*8,1))
vr = vr.reshape((31*8,1))

arreglo = np.concatenate((lodos, polimeros, torque, vr), axis=1)

####################################################################################################################
# ESCRITURA EXCEL
####################################################################################################################

# Se escribe en la cenda correspondiente del excel db_espesamiento_prueba.xlsx hoja valor_centrifuga

valores = excel_db["valor_centrifuga"]
celda_rf = 'E2'

matrix2excel(valores,celda_rf,arreglo)

excel_db.save(filename = excel_db_file)
